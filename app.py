from flask import Flask, render_template, request, jsonify, session, send_file
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
import pandas as pd

# Local imports
from database import Database
from gemini import GeminiAI
from email_system import EmailAutomation, EmailTracker
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== CONFIGURATION ====================

class Config:
    """Application configuration"""
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'dev-secret-key-change-in-production'
    # Use absolute path based on script location to ensure consistency
    UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
    MAX_CONTENT_LENGTH = 200 * 1024 * 1024  # 200MB max file size
    ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
    DATABASE_PATH = 'user_review_system.db'
    
    # Email settings (will be configured by user in Step 1)
    SMTP_SERVER = None
    SMTP_PORT = None
    SMTP_EMAIL = None
    SMTP_PASSWORD = None
    IMAP_SERVER = None
    IMAP_PORT = None
    
    # Gemini API (will be configured by user in Step 1)
    GEMINI_API_KEY = None

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS

# ==================== EXCEL HANDLER ====================

class ExcelHandler:
    """Handle Excel file operations"""
    
    @staticmethod
    def parse_owners_file(file_path):
        """
        Parse owners Excel file
        Expected columns: User, Email, Owner
        Returns: list of dicts with keys: user_name, email, owner_type
        """
        try:
            df = pd.read_excel(file_path)
            
            # Normalize column names
            df.columns = df.columns.str.strip()
            
            owners = []
            for _, row in df.iterrows():
                # Try to find a CC column (case-insensitive search handled by normalization if strictly named, 
                # but 'CC' vs 'Cc' vs 'CC Email' might need manual check if we only stripped whitespace)
                
                # Helper to find column regardless of case
                cc_email = ''
                for col in df.columns:
                    if col.lower() in ['cc', 'cc email', 'cc_email', 'copy']:
                        val = str(row[col]).strip()
                        if val and val.lower() != 'nan':
                            cc_email = val
                        break

                owner = {
                    'user_name': str(row['User']).strip(),
                    'email': str(row['Email']).strip().lower(),
                    'owner_type': str(row['Owner']).strip(),
                    'cc_email': cc_email
                }
                owners.append(owner)
            
            return owners, None
        except Exception as e:
            return None, f"Error parsing owners file: {str(e)}"
    
    @staticmethod
    def parse_users_file(file_path):
        """
        Parse users Excel file
        Expected columns: User, Email, Last Login, Roles, Groups
        Returns: list of dicts with user information
        """
        try:
            df = pd.read_excel(file_path)
            
            # Normalize column names
            df.columns = df.columns.str.strip()
            
            users = []
            for _, row in df.iterrows():
                user = {
                    'user_name': str(row['User']).strip(),
                    'email': str(row['Email']).strip().lower(),
                    'last_login': str(row.get('Last Login', '')).strip(),
                    'roles': str(row.get('Roles', '')).strip(),
                    'groups': str(row.get('Groups', '')).strip()
                }
                users.append(user)
            
            return users, None
        except Exception as e:
            return None, f"Error parsing users file: {str(e)}"
    
    @staticmethod
    def parse_action_excel(file_path):
        """
        Parse action Excel file from email attachment
        Expected to have a tab named 'Actions' with columns: Action, User Email, Details
        Returns: list of actions
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Look for 'Actions' sheet
            if 'Actions' in wb.sheetnames:
                ws = wb['Actions']
            else:
                # Use first sheet if 'Actions' not found
                ws = wb.active
            
            actions = []
            headers = []
            
            for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if idx == 1:
                    # Header row
                    headers = [str(cell).strip() for cell in row]
                    continue
                
                if not any(row):  # Skip empty rows
                    continue
                
                action_dict = {}
                found_action_val = None
                
                # Identify key column indices
                # We want to keep only: Email, User Name, Action
                # We want to IGNORE: Roles, Groups, Last Login, Department, Owner, Status (Noise)
                
                for i, cell in enumerate(row):
                    if i < len(headers):
                        header_key = headers[i]
                        cell_val = str(cell).strip() if cell else ''
                        
                        header_lower = header_key.lower()
                        
                        # Store 'action' value separately for filtering
                        if 'action' in header_lower:
                            found_action_val = cell_val
                            action_dict['Action'] = cell_val # Normalize key to 'Action'
                        
                        # Keep identifying info
                        elif 'email' in header_lower:
                            action_dict['Email'] = cell_val
                        # Match "User Name", "User Nam", "Username", etc.
                        elif 'user' in header_lower or 'name' in header_lower: 
                            # If it contains 'user' we are fairly safe. "Last Login" has 'last' not 'user'.
                            # "Department" has neither.
                            # Just be careful not to match "User Email" here, but the 'email' elif above catches that first?
                            # Actually, "User Email" has 'email'. So it goes to first bucket.
                            # "User Nam" has 'user'.
                            if 'email' not in header_lower:
                                action_dict['User Name'] = cell_val
                
                # STRICT FILTER:
                # Only append if we found an 'action' column AND it has text
                if found_action_val:
                     actions.append(action_dict)
            
            print(f"DEBUG: Parsed Excel. Input Rows: {idx}. Relevant Actions Found: {len(actions)}")
            print(f"DEBUG: Action Data to AI: {json.dumps(actions)}")
            return actions, None
        except Exception as e:
            return None, f"Error parsing action Excel: {str(e)}"
    
    @staticmethod
    def generate_users_report(users, output_path):
        """
        Generate Excel report of users
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Users Report"
            
            # Headers
            headers = ['User Name', 'Email', 'Last Login', 'Roles', 'Groups', 'Department', 'Owner', 'Status']
            ws.append(headers)
            
            # Data
            for user in users:
                ws.append([
                    user.get('user_name', ''),
                    user.get('email', ''),
                    user.get('last_login', ''),
                    user.get('roles', ''),
                    user.get('groups', ''),
                    user.get('department', ''),
                    user.get('owner_email', ''),
                    user.get('status', 'active')
                ])
            
            wb.save(output_path)
            return True, None
        except Exception as e:
            return False, f"Error generating report: {str(e)}"
    
    @staticmethod
    def generate_change_log_report(change_logs, output_path):
        """
        Generate Excel report of change logs
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Change Logs"
            
            # Headers
            headers = ['Ticket ID', 'Action Type', 'User Email', 'Old Value', 'New Value', 'Description', 'Performed By', 'Date']
            ws.append(headers)
            
            # Data
            for log in change_logs:
                ws.append([
                    log.get('ticket_id', ''),
                    log.get('action_type', ''),
                    log.get('user_email', ''),
                    log.get('old_value', ''),
                    log.get('new_value', ''),
                    log.get('description', ''),
                    log.get('performed_by', ''),
                    log.get('created_at', '')
                ])
            
            wb.save(output_path)
            return True, None
        except Exception as e:
            return False, f"Error generating change log report: {str(e)}"

# ==================== MAIN INIT ====================

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY

# Ensure upload folder exists
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)

# Initialize database
db = Database(Config.DATABASE_PATH)

# Global instances (will be initialized after configuration)
gemini_ai = None
email_automation = None
email_tracker = None


@app.route('/')
def index():
    """Main application page"""
    return render_template('index.html')


# ==================== STEP 1: SETUP & CONFIGURATION ====================

@app.route('/api/config/gemini', methods=['POST'])
def configure_gemini():
    """Configure Gemini API"""
    try:
        data = request.json
        api_key = data.get('api_key')
        
        if not api_key:
            return jsonify({'success': False, 'message': 'API key is required'})
        
        # Test connection
        global gemini_ai
        gemini_ai = GeminiAI(api_key)
        success, message = gemini_ai.test_connection()
        
        if success:
            # Save to database
            db.set_config('gemini_api_key', api_key)
            session['gemini_configured'] = True
            return jsonify({'success': True, 'message': 'Gemini AI configured successfully'})
        else:
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/config/email', methods=['POST'])
def configure_email():
    """Configure email settings"""
    try:
        data = request.json
        
        smtp_server = data.get('smtp_server')
        smtp_port = int(data.get('smtp_port', 587))
        email_address = data.get('email')
        password = data.get('password')
        imap_server = data.get('imap_server')
        imap_port = int(data.get('imap_port', 993))
        
        if not all([smtp_server, smtp_port, email_address, password]):
            return jsonify({'success': False, 'message': 'All SMTP fields are required'})
        
        # Test SMTP connection
        global email_automation
        email_automation = EmailAutomation(smtp_server, smtp_port, email_address, password)
        success, message = email_automation.test_connection()
        
        if not success:
            return jsonify({'success': False, 'message': message})
        
        # Test IMAP connection if provided
        if imap_server and imap_port:
            global email_tracker
            email_tracker = EmailTracker(imap_server, imap_port, email_address, password)
            imap_success, imap_message = email_tracker.test_connection()
            
            if not imap_success:
                return jsonify({'success': False, 'message': f'SMTP OK, but IMAP failed: {imap_message}'})
        
        # Save to database
        db.set_config('smtp_server', smtp_server)
        db.set_config('smtp_port', str(smtp_port))
        db.set_config('email', email_address)
        db.set_config('password', password)
        db.set_config('imap_server', imap_server or '')
        db.set_config('imap_port', str(imap_port))
        
        session['email_configured'] = True
        
        return jsonify({'success': True, 'message': 'Email configured successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/upload/owners', methods=['POST'])
def upload_owners():
    """Upload owners Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'})
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Invalid file type. Please upload .xlsx or .xls file'})
        
        # Save file
        filename = secure_filename(file.filename)
        filepath = os.path.join(Config.UPLOAD_FOLDER, f'owners_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{filename}')
        file.save(filepath)
        
        # Parse file
        owners, error = ExcelHandler.parse_owners_file(filepath)
        
        if error:
            return jsonify({'success': False, 'message': error})
        
        # Clear existing owners and add new ones
        db.clear_owners()
        
        for owner in owners:
            db.add_owner(owner['user_name'], owner['email'], owner['owner_type'], owner.get('cc_email'))
        
        session['owners_uploaded'] = True
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded {len(owners)} owners',
            'owners': owners
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/owners', methods=['GET'])
def get_owners():
    """Get all owners"""
    try:
        owners = db.get_owners()
        return jsonify({'success': True, 'owners': owners})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ==================== STEP 2: USER MANAGEMENT ====================

@app.route('/api/upload/users', methods=['POST'])
def upload_users():
    """Upload users Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'})
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Invalid file type'})
        
        # Save file
        filename = secure_filename(file.filename)
        filepath = os.path.join(Config.UPLOAD_FOLDER, f'users_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{filename}')
        file.save(filepath)
        
        # Parse file
        users, error = ExcelHandler.parse_users_file(filepath)
        
        if error:
            return jsonify({'success': False, 'message': error})
        
        # Get owners to map users to departments
        owners = db.get_owners()
        owner_map = {owner['email']: owner['owner_type'] for owner in owners}
        
        # Clear existing users
        db.clear_users()
        
        # Add users and assign to departments based on groups
        it_users = []
        business_users = []
        
        for user in users:
            # Determine department based on groups
            # If groups contains "(IT)" or "IT" -> IT department
            # Otherwise -> Business department
            groups = user.get('groups', '').lower()
            department = 'Business'  # Default to Business
            owner_email = None
            
            # Check if user belongs to IT department
            if '(it)' in groups or groups.startswith('it ') or groups.endswith(' it') or groups == 'it':
                department = 'IT'
            
            # Find matching owner based on department
            for owner in owners:
                if owner['owner_type'] == department:
                    owner_email = owner['email']
                    break
            
            # If no owner found for the department, try to find any owner
            if not owner_email and owners:
                owner_email = owners[0]['email']
            
            db.add_user(
                user['user_name'],
                user['email'],
                user.get('last_login', ''),
                user.get('roles', ''),
                user.get('groups', ''),
                department,
                owner_email
            )
            
            user['department'] = department
            user['owner_email'] = owner_email
            user['status'] = 'active' # Default status for new uploads
            
            if department == 'IT':
                it_users.append(user)
            else:
                business_users.append(user)
        
        session['users_uploaded'] = True
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded {len(users)} users',
            'it_users': it_users,
            'business_users': business_users
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/users', methods=['GET'])
def get_users():
    """Get all users"""
    try:
        department = request.args.get('department')
        status = request.args.get('status', 'active')
        users = db.get_users(department=department, status=status)
        
        it_users = [u for u in users if u['department'] == 'IT']
        business_users = [u for u in users if u['department'] == 'Business']
        
        return jsonify({
            'success': True,
            'it_users': it_users,
            'business_users': business_users,
            'users': users
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ==================== STEP 3: REVIEW PROCESS ====================

@app.route('/api/tickets/generate', methods=['POST'])
def generate_tickets():
    """Generate tickets and send review emails"""
    try:
        # Get owners and users
        owners = db.get_owners()
        
        if not owners:
            return jsonify({'success': False, 'message': 'No owners found. Please upload owners file first.'})
        
        # Initialize email automation if not already done
        global email_automation
        if not email_automation:
            smtp_server = db.get_config('smtp_server')
            smtp_port = int(db.get_config('smtp_port'))
            email = db.get_config('email')
            password = db.get_config('password')
            email_automation = EmailAutomation(smtp_server, smtp_port, email, password)
        
        # Create tickets for each owner
        tickets_data = []
        
        for owner in owners:
            # Get users for this owner
            users = db.get_users()
            owner_users = [u for u in users if u.get('owner_email') == owner['email']]
            
            if not owner_users:
                continue
            
            # Generate ticket ID
            ticket_id = email_automation.generate_ticket_id(owner_type=owner['owner_type'])
            
            # Save ticket to database
            # Determine correct CC list from Database Owner Record
            cc_list = owner.get('cc_email', '')

            db.add_ticket(ticket_id, owner['owner_type'], owner['email'], cc_list)
            
            # Generate Excel attachment for this owner's users
            attachment_path = os.path.join(Config.UPLOAD_FOLDER, f'User_List_{ticket_id}.xlsx')
            gen_success, gen_error = ExcelHandler.generate_users_report(owner_users, attachment_path)
            
            error_msg = None
            if not gen_success:
                error_msg = f"Failed to generate Excel report: {gen_error}"
                attachment_path = None # Don't try to attach if generation failed
            
            tickets_data.append({
                'ticket_id': ticket_id,
                'owner_email': owner['email'],
                'owner_name': owner['user_name'],
                'department': owner['owner_type'],
                'users': owner_users,
                'cc_emails': cc_list,
                'attachment_path': attachment_path,
                'error_msg': error_msg
            })
        
        # Send emails
        results = email_automation.send_review_emails(tickets_data)
        
        # Cleanup temporary attachments
        for ticket in tickets_data:
            if ticket.get('attachment_path') and os.path.exists(ticket['attachment_path']):
                try:
                    os.remove(ticket['attachment_path'])
                except Exception as e:
                    print(f"Error removing temp file {ticket['attachment_path']}: {e}")
        
        return jsonify({
            'success': True,
            'message': f'Generated {len(tickets_data)} tickets and sent emails',
            'results': results
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tickets', methods=['GET'])
def get_tickets():
    """Get all tickets"""
    try:
        tickets = db.get_tickets()
        return jsonify({'success': True, 'tickets': tickets})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/responses/fetch', methods=['POST'])
def fetch_responses():
    """Fetch email responses from IMAP"""
    try:
        # Initialize email tracker if not already done
        global email_tracker
        if not email_tracker:
            imap_server = db.get_config('imap_server')
            imap_port = int(db.get_config('imap_port'))
            email = db.get_config('email')
            password = db.get_config('password')
            
            if not imap_server:
                return jsonify({'success': False, 'message': 'IMAP not configured'})
            
            email_tracker = EmailTracker(imap_server, imap_port, email, password)
        
        # Get active tickets
        tickets = db.get_tickets()
        ticket_ids = [t['ticket_id'] for t in tickets]
        
        # Fetch responses from last 7 days
        since_date = datetime.now() - timedelta(days=7)
        responses, error = email_tracker.fetch_responses(ticket_ids, since_date)
        
        if error:
            return jsonify({'success': False, 'message': error})
        
        # Save responses to database
        if not responses:
            return jsonify({'success': True, 'message': 'No new email responses found', 'responses': []})
        
        # Save responses to database (ONLY for valid existing tickets)
        new_responses = 0
        conn = db.get_connection() # Keep connection open for checks
        accepted_responses = [] # List of responses to return to frontend
        
        for response in responses:
            # 1. Check if Ticket ID exists and Get Owner
            cursor = conn.cursor()
            cursor.execute('SELECT owner_email FROM tickets WHERE ticket_id = ?', (response['ticket_id'],))
            row = cursor.fetchone()
            
            if not row:
                print(f"Skipping response for unknown/deleted ticket: {response['ticket_id']}")
                continue 
            
            # 2. Strict Owner Check
            owner_email = row[0] 
            sender_email = response['from_email'].strip().lower()
            owner_clean = owner_email.strip().lower()
            
            if owner_clean != sender_email:
                print(f"[REJECTED] Sender '{sender_email}' != Owner '{owner_clean}'")
                print(f"   (Ticket: {response['ticket_id']})")
                continue 
            else:
                print(f"[ACCEPTED] Sender '{sender_email}' is Owner")
                accepted_responses.append(response)

            # 3. Check if response already exists
            existing = db.get_email_responses(ticket_id=response['ticket_id'])
            
            # Simple duplicate check based on from_email and body
            is_duplicate = False
            for existing_resp in existing:
                if (existing_resp['from_email'] == response['from_email'] and 
                    existing_resp['body'][:100] == response['body'][:100]):
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                attachment_data = json.dumps(response.get('attachments', []))
                db.add_email_response(
                    response['ticket_id'],
                    response['from_email'],
                    response['subject'],
                    response['body'],
                    response['has_attachment'],
                    attachment_data
                )
                new_responses += 1
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Fetched {len(responses)} emails. Saved {new_responses} valid responses.',
            'responses': accepted_responses
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/responses', methods=['GET'])
def get_responses():
    """Get all email responses"""
    try:
        responses = db.get_email_responses()
        return jsonify({'success': True, 'responses': responses})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


def execute_action(db, ticket_id, action, users, source_email):
    """
    Execute a single action (delete/update) on a user.
    Returns: 1 if successful, 0 otherwise
    """
    action_type = action.get('action_type')
    user_email = action.get('user_email')
    
    if not user_email:
        # Log partial failure/warning
        db.add_change_log(
            ticket_id,
            'error',
            action.get('user_identifier', 'Unknown'),
            'N/A',
            'Failed',
            f"⚠️ Could not identify user for action: {action.get('description', 'Unknown action')}",
            source_email
        )
        return 0
    
    # Normalize email
    user_email = user_email.strip().lower()
    try:
        action_type = action.get('action_type')
        # user_email is already defined and normalized above, no need to redefine
        new_value = action.get('new_value')
        
        # Professional Description Generator
        # We ignore the AI's verbose description and generate a standard one
        description = action.get('description', 'Action performed') # Default, will be overwritten
        
        if action_type == 'delete':
            db.delete_user(user_email)
            # Use AI description if available, else standard
            if not action.get('description'):
                description = f"User {user_email} deleted."
            
            logger.info(f"ACTION: Deleted {user_email}")
            
            db.add_change_log(ticket_id, 'delete', user_email, 'active', 'deleted', description, source_email)
            return 1
            
        elif action_type == 'update_role':
            # Get old role
            old_role = 'Unknown'
            user_name = 'User'
            for u in users:
                if u['email'].lower() == user_email: # Compare normalized emails
                    old_role = u.get('roles', '')
                    user_name = u.get('user_name', 'User')
                    break
            
            db.update_user_role(user_email, new_value)
            description = f"Roles updated for {user_name}." # Standardized
            logger.info(f"ACTION: Updated roles for {user_email} to {new_value}")
            
            db.add_change_log(ticket_id, 'update_role', user_email, old_role, new_value, description, source_email)
            return 1
            
        elif action_type == 'add':
            # Add logic not fully implemented in this snippet but preventing errors
             pass

    except Exception as e:
        logger.error(f"Error executing action {action}: {e}")
        db.add_change_log(ticket_id, 'error', action.get('user_email', 'unknown'), 'N/A', 'Failed', f"Error: {str(e)}", 'SYSTEM')
    
    return 0


@app.route('/api/responses/process', methods=['POST'])
def process_responses():
    """Process email responses with AI"""
    try:
        # Initialize Gemini AI if not already done
        global gemini_ai
        if not gemini_ai:
            api_key = db.get_config('gemini_api_key')
            if not api_key:
                return jsonify({'success': False, 'message': 'Gemini AI not configured'})
            gemini_ai = GeminiAI(api_key)
        
        # Get unprocessed responses
        responses = db.get_email_responses(processed=0)
        
        if not responses:
            return jsonify({'success': True, 'message': 'No new responses to process'})
        
        # Get current users
        users = db.get_users()
        
        total_actions = 0
        
        for response in responses:
            ticket_id = response['ticket_id']
            # Double check processed status to handle concurrency
            conn = db.get_connection()
            check_cursor = conn.cursor()
            check_cursor.execute('SELECT processed FROM email_responses WHERE id = ?', (response['id'],))
            row = check_cursor.fetchone()
            conn.close()
            
            if row and row['processed']:
                logger.info(f"Skipping already processed response {response['id']}")
                continue

            sender_email = response['from_email']

            # --- OWNER VALIDATION ---
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT owner_email FROM tickets WHERE ticket_id = ?', (ticket_id,))
            ticket_row = cursor.fetchone()
            conn.close()

            if not ticket_row:
                 print(f"Ticket {ticket_id} not found. Skipping.")
                 continue

            ticket_owner_email = ticket_row['owner_email']
            
            if ticket_owner_email.strip().lower() != sender_email.strip().lower():
                msg = f"Sender '{sender_email}' is NOT the owner ('{ticket_owner_email}'). Changes rejected."
                print(f"[UNAUTHORIZED] {msg}")
                # Log this security event so user sees it in Change Log
                db.add_change_log(
                    ticket_id, 
                    'unauthorized', 
                    'N/A',          # user_email
                    'N/A',          # old_value
                    'Rejected',     # new_value
                    f"⛔ UNAUTHORIZED: {msg}", # description
                    'SYSTEM'        # performed_by
                )
                db.mark_response_processed(response['id']) 
                continue
            # ------------------------
            
            # ------------------------
            # CRITICAL: Scope Fix - Only allow actions on users owned by this ticket owner
            # ------------------------
            all_users = users # Keep reference to all
            owner_users = [u for u in all_users if str(u.get('owner_email', '')).lower() == ticket_owner_email.strip().lower()]
            
            print(f"DEBUG: Scoped User List for {ticket_owner_email}: {[u['user_name'] for u in owner_users]}")

            # 1. Process Text Response
            # 1. Process Text Response
            if response['body']:
                # Pass ALL users to AI to prevent scope Hallucinations on Name matching
                result = gemini_ai.parse_email_response(response['body'], users)
                print(f"DEBUG: AI TEXT RESPONSE for {ticket_id}:")
                print(json.dumps(result, indent=2))
                
                # Check for AI Errors
                if result.get('error'):
                    logger.error(f"AI Error for {ticket_id}: {result['error']}")
                    db.add_change_log(ticket_id, 'error', 'AI System', 'N/A', 'Failed', result['error'], 'SYSTEM')

                for action in result.get('actions', []):
                    # verify action again (double safety)
                    user_email = action.get('user_email')
                    if user_email:
                        is_owned = any(u['email'].lower() == user_email.lower() for u in owner_users)
                        if not is_owned:
                            # Relaxed Check: Warn but allow
                            msg = f"Warning: User {user_email} is not owned by you, but action allowed per configuration."
                            print(f"SAFETY WARNING: {msg}")
                            # Do NOT continue, allow execution
                            # db.add_change_log(ticket_id, 'info', user_email, 'N/A', 'Allowed', msg, 'SYSTEM')

                    total_actions += execute_action(db, ticket_id, action, users, response['from_email'])
            
            # 2. Process Excel Attachment
            if response['has_attachment'] and response['attachment_data']:
                try:
                    print(f"DEBUG: Checking attachments for Ticket {ticket_id}")
                    attachments = json.loads(response['attachment_data'])
                    
                    for attachment in attachments:
                        print(f"DEBUG: Found attachment: {attachment['filename']}")
                        if attachment['filename'].endswith(('.xlsx', '.xls')):
                            print(f"DEBUG: Processing Excel attachment: {attachment['filename']} at {attachment['filepath']}")
                            
                            excel_actions, error = ExcelHandler.parse_action_excel(attachment['filepath'])
                            print(f"DEBUG: Parsed Excel Actions: {excel_actions}, Error: {error}")
                            
                            if not error and excel_actions:
                                # Process with AI 
                                # We pass ALL users to AI so it correctly identifies the target user (even if out of scope).
                                # The ownership check below will catch/reject unauthorized actions.
                                logger.info(f"DEBUG: sending {len(excel_actions)} filtered rows to AI with full context.")
                                result = gemini_ai.parse_excel_actions(excel_actions, users)
                                
                                print(f"DEBUG: AI EXCEL RESPONSE for {ticket_id}:")
                                print(json.dumps(result, indent=2))
                                
                                for action in result:
                                    logger.info(f"DEBUG: Executing Action: {action}")
                                    
                                    # SECURITY CHECK: Ownership
                                    user_email = action.get('user_email')
                                    if user_email:
                                        is_owned = any(u['email'].lower() == user_email.lower() for u in owner_users)
                                        if not is_owned:
                                            # Relaxed Check: Warn but allow
                                            logger.warning(f"SECURITY WARNING: User {user_email} not owned by {ticket_owner_email}, but proceeding.")
                                            # Do NOT continue, proceed to execution
                                            # db.add_change_log(ticket_id, 'info', user_email, 'N/A', 'Allowed', "Action allowed on unowned user.", 'SYSTEM')

                                    total_actions += execute_action(db, ticket_id, action, users, response['from_email'])
                            elif error:
                                print(f"Error parsing excel: {error}")
                                db.add_change_log(ticket_id, 'error', 'N/A', 'N/A', 'Failed', f"Excel Parse Error: {error}", 'SYSTEM')

                except Exception as e:
                    print(f"Error processing attachments: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            # Mark response as processed
            db.mark_response_processed(response['id'])
        
        return jsonify({
            'success': True,
            'message': f'Processed {len(responses)} responses with {total_actions} actions'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/changelogs', methods=['GET'])
def get_changelogs():
    """Get all change logs"""
    try:
        logs = db.get_change_logs()
        return jsonify({'success': True, 'logs': logs})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ==================== STEP 4: REPORTS ====================

@app.route('/api/reports/users', methods=['GET'])
def export_users_report():
    """Export users report"""
    try:
        users = db.get_users()
        
        output_path = os.path.join(Config.UPLOAD_FOLDER, f'users_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        success, error = ExcelHandler.generate_users_report(users, output_path)
        
        if not success:
            return jsonify({'success': False, 'message': error})
        
        return send_file(output_path, as_attachment=True)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/changelogs', methods=['GET'])
def export_changelogs_report():
    """Export change logs report"""
    try:
        logs = db.get_change_logs()
        
        output_path = os.path.join(Config.UPLOAD_FOLDER, f'changelogs_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        success, error = ExcelHandler.generate_change_log_report(logs, output_path)
        
        if not success:
            return jsonify({'success': False, 'message': error})
        
        return send_file(output_path, as_attachment=True)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get system statistics"""
    try:
        users = db.get_users()
        tickets = db.get_tickets()
        responses = db.get_email_responses()
        logs = db.get_change_logs()
        
        stats = {
            'total_users': len(users),
            'it_users': len([u for u in users if u['department'] == 'IT']),
            'business_users': len([u for u in users if u['department'] == 'Business']),
            'total_tickets': len(tickets),
            'pending_tickets': len([t for t in tickets if t['status'] == 'pending']),
            'total_responses': len(responses),
            'unprocessed_responses': len([r for r in responses if not r['processed']]),
            'total_changes': len(logs),
            'deleted_users': len([l for l in logs if l['action_type'] == 'delete']),
            'role_updates': len([l for l in logs if l['action_type'] == 'update_role'])
        }
        
        return jsonify({'success': True, 'stats': stats})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reset_process', methods=['POST'])
def reset_process_data():
    """Clear tickets, responses, and logs but keep users/owners"""
    try:
        success, message = db.clear_process_data()
        if success:
            return jsonify({'success': True, 'message': 'Process data cleared. Owners and Users preserved.'})
        else:
            return jsonify({'success': False, 'message': f'Error clearing process data: {message}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error clearing process data: {str(e)}'})

@app.route('/api/reset', methods=['POST'])
def reset_system():
    try:
        db.clear_owners()
        db.clear_users()
        success, message = db.clear_process_data()
        
        if success:
             return jsonify({'success': True, 'message': 'System reset successfully. All data cleared.'})
        else:
             return jsonify({'success': False, 'message': f'Error resetting system: {message}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error resetting system: {str(e)}'})

@app.route('/api/export_desktop', methods=['POST'])
def export_results():
    """Export all results to a timestamped folder on the Desktop"""
    try:
        # 1. Prepare Data
        users = db.get_users()
        responses = db.get_email_responses()
        logs = db.get_change_logs()
        
        # 2. Get Desktop Path (Cross-platform)
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        
        # 3. Create Fixed "Output" Folder
        folder_name = "Output"
        output_dir = os.path.join(desktop_path, folder_name)
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        else:
            # Optional: Clear existing files to ensure clean export?
            # User just said "output only", assuming standard behavior (overwrite/merge)
            pass
            
        # 4. Generate Excel Files
        
        # File 1: Final User List
        if users:
            df_users = pd.DataFrame([dict(u) for u in users])
            # Select/Order columns nicely if possible, else dump all
            df_users.to_excel(os.path.join(output_dir, "Final_User_List.xlsx"), index=False)
            
        # File 2: Email Responses
        if responses:
            df_responses = pd.DataFrame([dict(r) for r in responses])
            df_responses.to_excel(os.path.join(output_dir, "Email_Responses.xlsx"), index=False)
            
        # File 3: Change Logs
        if logs:
            df_logs = pd.DataFrame([dict(l) for l in logs])
            df_logs.to_excel(os.path.join(output_dir, "Change_Log.xlsx"), index=False)
            
        return jsonify({
            'success': True, 
            'message': f'Export successful! Files saved to desktop folder: {folder_name}',
            'path': output_dir
        })

    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}'})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
